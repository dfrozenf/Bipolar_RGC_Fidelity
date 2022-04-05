#!/usr/bin/env python3
import os
import re
from enum import Enum, auto
from itertools import groupby
from itertools import zip_longest
import math
from collections import OrderedDict
from scipy import stats
import numpy as np
from openpyxl import Workbook

"""
Directory wherein all experimental data is stored. Can be recursively organized.
"""
LOCALDIR = 'Data/'

"""
Bin size for latency frequency distributions.
"""
LATENCYSTEP = 0.1


class Presets(Enum):
    NIGHT_1 = auto()
    NIGHT_2 = auto()
    NIGHT_3 = auto()
    NIGHT_4 = auto()
    CONTRAST = auto()
    SPATIAL = auto()


"""Experiment-specific values"""
TIMEOUTS = {Presets.NIGHT_3: 30, Presets.NIGHT_4: 10, Presets.CONTRAST: 10, Presets.SPATIAL: 10}
CONTRAST_LVLS = {1: 1, 2: 2, 4: 4, 7: 8, 14: 16, 27: 32, 52: 64, 100: 100}


class ImageTypes(Enum):
    REWARD = "REWARD"
    CONTROL = "CONTROL"


class DoorStates(Enum):
    High, Low = auto(), auto()


class PumpStates(Enum):
    On, Off = auto(), auto()


class Activity(Enum):
    Running, Poking = auto(), auto()


class Mouse:

    def __init__(self, cageNum, rotation_intervals, poke_events):
        self._cageNum = cageNum
        self._rotation_intervals = rotation_intervals
        self._poke_events = poke_events

    @property
    def rotation_intervals(self):
        return self._rotation_intervals

    @property
    def poke_events(self):
        return self._poke_events


class RotationInterval:
    # contiguous series of wheel spins

    def __init__(self, halfTimes, image):
        self._halfTimes = []
        self._image = image
        self.viable = True
        self._rpms = []  # instantaneous speeds in rotations per minute
        raw_rpms = []  # prone to error, will be refined
        for i in range(1, len(halfTimes) - 1):
            raw_rpms.append(60 / (halfTimes[i + 1] - halfTimes[i - 1]))
            # instantaneous speeds at beginning and end of rotation event
        raw_rpms.insert(0, 30 / (halfTimes[1] - halfTimes[0]))
        raw_rpms.append(
            30 / (halfTimes[-1] - halfTimes[-2]))  # instantaneous speeds at beginning and end of rotation event
        erratic = []
        for i in range(1, len(halfTimes) - 1):
            if raw_rpms[i] > 200:
                erratic.append(i)
        for i in range(1, len(halfTimes) - 1):
            if i not in erratic:
                self._rpms.append(raw_rpms[i])
                self._halfTimes.append(halfTimes[i])
        if len(self._halfTimes) < 2:
            self.viable = False

    def numRotations(self):
        return len(self._halfTimes) // 2

    def avgSpeed(self):
        # average speed in RPM
        return self.numRotations() * 60 / (self._halfTimes[-1] - self._halfTimes[0])

    @property
    def speeds(self):
        return self._rpms

    @property
    def startTime(self):
        return self._halfTimes[0]

    @property
    def midTime(self):
        return (self._halfTimes[-1] + self.halfTimes[0]) / 2

    @property
    def halfTimes(self):
        return self._halfTimes

    @property
    def image(self):
        return self._image


class PokeEvent:
    # series of repeated pokes

    def __init__(self, doorStates, doorTimes, pumpTimes, pumpStates, image):
        self._doorStates = doorStates
        self._doorTimes = doorTimes
        self._pumpTimes = pumpTimes
        self._pumpStates = pumpStates
        self._image = image
        self._imageAppearanceTime = image.latestAppearanceTime()
        self._imageAppearance = self._image.appearances.get(self._imageAppearanceTime)
        self._imageAppearance.addPokeEvent(self)
        # add this pokeEvent to the image appearance during which it occured
        s, t = self.successfulPokes()
        if s == 1:
            self.latency = t[0] - self._imageAppearanceTime
            self.pokeTime = t[0]
        else:
            self.latency = None
            self.pokeTime = None

    def isSuccess(self):
        successful = False
        for p in self._pumpStates:
            if p is PumpStates.On:
                successful = True
        return successful

    def successfulPokes(self):
        num = 0
        times = []
        for p, t in zip(self._pumpStates, self._pumpTimes):
            if p is PumpStates.On:
                num += 1
                times.append(t - 0.003)  # Pump is activated 3 ms after poke occurs
        return num, times

    def allPokes(self):
        num = 0
        times = []
        for p, t in zip(self._doorStates, self._doorTimes):
            if p is DoorStates.Low:
                num += 1
                times.append(t)
        return num, times

    def unsuccessfulPokes(self):
        successfulPokes = set(self.successfulPokes()[1])
        allPokes = set(self.allPokes()[1])
        unsuccessful = list(allPokes - successfulPokes)
        unsuccessful.sort()
        return len(unsuccessful), unsuccessful

    def totalPokesNoTimeout(self, grace=30):
        # returns total number of pokes EXCLUDING those that are failed due to image timeout
        # i.e. after pump success
        critical_time = None
        for p, t in zip(self._pumpStates, self._pumpTimes):
            if p is PumpStates.On:
                critical_time = t
        if critical_time is None or self.successfulPokes()[0] > 1:
            return self.allPokes()[0]
        else:
            beforeSuccessful = 0
            for dt in self._doorTimes:
                if dt <= critical_time or dt > critical_time + grace:  # 30 seconds grace period
                    beforeSuccessful += 1
            return int(math.ceil(beforeSuccessful / 2))
            # two door states constitute a full poke
            # ceil necessary because only door opening documented before poke

    def drinkTimes(self):
        drinkStart = 0
        drinkTimes = []
        for p, t in zip(self._pumpStates, self._pumpTimes):
            if p is PumpStates.On:
                drinkStart = t
            else:
                drinkTimes.append(t - drinkStart)
        return drinkTimes

    @property
    def startTime(self):
        return self._doorTimes[0]

    @property
    def image(self):
        return self._image

    @property
    def doorTimes(self):
        return self._doorTimes

    @property
    def doorStates(self):
        return self._doorStates

    @property
    def pumpTimes(self):
        return self._pumpTimes

    @property
    def pumpStates(self):
        return self._pumpStates

    @property
    def imageAppearanceTime(self):
        return self._imageAppearanceTime

    @property
    def imageAppearance(self):
        return self._imageAppearance


class Appearance:

    def __init__(self, image, time, old_img):
        self._time = time
        self._image = image
        self._poke_events = []
        if old_img.appearances:
            old_app_num = old_img.appearances.get(old_img.latestAppearanceTime()).rewardSeqNum
            self._rewardSeqNum = 0 if image.imageType == ImageTypes.CONTROL else old_app_num + 1
        else:
            self._rewardSeqNum = 0 if image.imageType == ImageTypes.CONTROL else 1

    def addPokeEvent(self, poke_event):
        self._poke_events.append(poke_event)

    @property
    def poke_events(self):
        return self._poke_events

    @property
    def time(self):
        return self._time

    @property
    def image(self):
        return self._image

    @property
    def rewardSeqNum(self):
        return self._rewardSeqNum


class Image:
    """
    Log of image appearances by time.
    """
    appearanceLog = OrderedDict()

    def __init__(self, name, imageType):
        self.name = name
        assert isinstance(imageType, ImageTypes), 'use ImageType enum to assign images'
        self.imageType = imageType
        self._appearanceTimes = []
        self._appearances = {}
        if imageType == ImageTypes.REWARD:
            self.avg_latency = None
            self.SEM_latency = None

    def __eq__(self, other):
        return self.name == other.name and self.imageType == other.imageType

    def __hash__(self):
        return hash(self.name)

    def incrementAppearances(self, time, old_img):
        self._appearanceTimes.append(time)
        self._appearances[time] = Image.appearanceLog[time] = Appearance(self, time, old_img)

    @property
    def numAppearances(self):
        return len(self._appearanceTimes)

    @property
    def appearanceTimes(self):
        return self._appearanceTimes

    def latestAppearanceTime(self):
        return self._appearanceTimes[-1]

    @property
    def appearances(self):
        return self._appearances

    @staticmethod
    def imageAtTime(time):
        appearances = list(Image.appearanceLog.keys())
        return Image.appearanceLog[max(filter(lambda k: k < time, appearances))].image


# def cumulativeSuccess(poke_events):
#     outcomes = [int(pe.isSuccess()) for pe in poke_events]
#     print("Successful Poke Events: {0}".format(sum(outcomes)))
#     # times = [pe.startTime for pe in poke_events]
#     cumulative_success = 0
#     total = 0
#     cumulative_probabilities = []
#     for outcome in outcomes:
#         cumulative_success += outcome
#         total += 1
#         cumulative_probabilities.append(cumulative_success / total)
#     plt.plot(cumulative_probabilities, marker='.')
#     plt.ylabel('Cumulative Probability')
#     plt.xlabel('Poke Events')
#     plt.title('Poke Success Rate')
#     plt.show()

def setImageLatencies(poke_events, images):
    pe_by_imageAppTime = {}  # poke events per image appearance times
    for pe in poke_events:
        if pe.latency is not None:
            pe_by_imageAppTime[pe.imageAppearanceTime] = pe
    rewardImgs = filter(lambda im: im.imageType is ImageTypes.REWARD, images)
    for ri in rewardImgs:
        img_latencies = []
        img_latencies_1st = []
        for img_app_time in ri.appearanceTimes:
            pe = pe_by_imageAppTime.get(img_app_time, None)
            if pe is not None:
                img_latencies.append(pe.latency)
                if ri.appearances.get(img_app_time).rewardSeqNum == 1:
                    img_latencies_1st.append(pe.latency)
        ri.avg_latency = np.mean(img_latencies)
        ri.avg_latency_1st = np.mean(img_latencies_1st)
        ri.SEM_latency = stats.sem(img_latencies)
        ri.SEM_latency_1st = stats.sem(img_latencies_1st)


def getContrast(image):

    if "negative" in image.name.lower():
        return 0

    try:
        contrastVals = [int(s) for s in image.name if s.isdigit()]
        contrastVal = int(''.join(str(digit) for digit in contrastVals))
    except (IndexError, ValueError):
        contrastVal = 100  # if not level specified, assume 100
    return CONTRAST_LVLS.get(contrastVal, contrastVal)


def pokeLatencies(wb, preset):
    outputCSV = wb.active
    allLatencies = []
    trueLatencies = []
    rewardTimes = []
    imageWiseLatencies = {}
    outputCSV.append([])
    outputCSV.append(['Time of REWARD', 'Image Contrast Level', 'Latencies (sec)'])

    ## contrast instead of name
    ### add time in hours
    for ap in Image.appearanceLog.values():
        if ap.image.imageType != ImageTypes.REWARD:
            continue
        elif ap.rewardSeqNum != 1:  # only first appearances should be considered
            continue

        contrastLevel = getContrast(ap.image)

        if not ap.poke_events:
            outputCSV.append([ap.time, contrastLevel, TIMEOUTS.get(preset)])
            allLatencies.append(TIMEOUTS.get(preset))
            rewardTimes.append(ap.time)

        else:
            if imageWiseLatencies.get(ap.image) is None:
                imageWiseLatencies[ap.image] = []
            for pe in ap.poke_events:
                if pe.latency is not None:
                    outputCSV.append([pe.imageAppearanceTime, contrastLevel, pe.latency])
                    allLatencies.append(pe.latency)
                    trueLatencies.append(pe.latency)
                    rewardTimes.append(pe.imageAppearanceTime)
                    imageWiseLatencies[ap.image].append(pe.latency)
                # NOTE that a poke event has a LATENCY of NONE iff the poke was unsuccessful.
                # Because latencies are considered only for reward images and REWARD images are
                # reset once the first poke ceases, such a case shall not be encountered unless
                # an erroneous wheel rotation causes event switching and falsely creates two events
                # one successful and the other unsuccessful.

    if len(trueLatencies) == 0:
        return

    ws2 = wb.create_sheet(title='All')
    headings = ["Time", "Latency", ""]
    sheetData = [[r / (60 ** 2) for r in rewardTimes], allLatencies, []]  # time in hours
    ws2.append(headings)
    for row in zip_longest(*sheetData, fillvalue=""):
        try:
            ws2.append(row)
        except ValueError:
            pass

    sheetData = []
    headings = []
    ws3 = wb.create_sheet(title='Image-wise')
    for im in sorted(imageWiseLatencies.keys(), key=getContrast):
        headings.extend(["Image Contrast", "Latency"])
        headings.append("")
        latencies = imageWiseLatencies.get(im)
        sheetData.append([getContrast(im)] * len(latencies) + ["", "MEAN", "SEM"])
        sheetData.append(latencies + ["", im.avg_latency, im.SEM_latency])
        sheetData.append([])
    ws3.append(headings)
    for row in zip_longest(*sheetData, fillvalue=""):
        try:
            ws3.append(row)
        except ValueError:
            pass

    sheetData = []
    headings = []
    ws4 = wb.create_sheet(title='Distributions')
    for im in sorted(imageWiseLatencies.keys(), key=getContrast):
        headings.extend(["Contrast", "Bin", "Count", "Rel. Frequency", ""])
        latencies = imageWiseLatencies.get(im)
        count, hbin = np.histogram(latencies, bins=np.arange(0, TIMEOUTS.get(preset, 10) + LATENCYSTEP, LATENCYSTEP))
        count = list(count)
        hbin = list(hbin)
        sheetData.append([getContrast(im)] * len(hbin))
        sheetData.append(hbin + ["", "Total"])
        sheetData.append(count + ["", "", sum(count)])
        percents = [c * 100 / len(latencies) for c in count]
        sheetData.append(percents + ["", "", sum(percents)])  # relative frequencies as %
        sheetData.append([])
        # sheetData = list(map(lambda arr: list(map(lambda k: str(k), arr)), sheetData))
    ws4.append(headings)
    for row in zip_longest(*sheetData, fillvalue=""):
        try:
            ws4.append(row)
        except ValueError:
            pass


def pokesPerHour(poke_events, outputCSV):
    hourlyPokes = {}  # dictionary stores pokes for each hour
    for pe in poke_events:
        for t, s in zip(pe.pumpTimes, pe.pumpStates):
            if s is PumpStates.On:
                hr = int(t / 3600) + 1  # convert t to hours, round up for nth hour
                hourlyPokes[hr] = hourlyPokes.get(hr,
                                                  0) + 1  # increment pokes for each hour, default value of 0 supplied to initialize
    outputCSV.append([])
    outputCSV.append(['Hour', '# Successful Pokes'])
    for k in range(1, 13):
        print("Successful pokes in hour #{0} >> {1}".format(k, hourlyPokes.get(k, 0)))
        outputCSV.append([k, hourlyPokes.get(k, 0)])


# def numPokes(poke_events):
#     allPokes = [pe.allPokes()[0] for pe in poke_events]
#     plt.hist(allPokes)
#     plt.xlabel("Number of pokes per poke event")
#     plt.ylabel("Frequency")
#     plt.show()
#
#
# def drinkLengths(poke_events):
#     lengths = []
#     for pe in poke_events:
#         lengths.extend(pe.drinkTimes())
#     plt.hist(lengths)
#     plt.xlabel("Time (sec) drinking sugar water")
#     plt.ylabel("Frequency")
#     plt.show()


def endRun(wheelHalfTimes, image, rotation_intervals):
    if len(wheelHalfTimes) < 3:
        return
    rotation_intervals.append(RotationInterval(wheelHalfTimes, image))  # add this interval to list


def endPoke(doorStates, doorTimes, pumpTimes, pumpStates, image, poke_events):
    poke_events.append(PokeEvent(doorStates, doorTimes, pumpTimes, pumpStates, image))


def pruneRotationIntervals(rotation_intervals):
    erratic = []
    for ri in rotation_intervals:
        if not ri.viable:
            erratic.append(ri)
    for ri in erratic:
        rotation_intervals.remove(ri)


def pokeStatistics(poke_events, images, filename, outputCSV, preset):
    successful = 0
    total = 0
    i = 0
    for pe in poke_events:
        i += 1
        successful += pe.successfulPokes()[0]
        total += pe.totalPokesNoTimeout()
    rewardImgs = list(filter(lambda im: im.imageType is ImageTypes.REWARD, images))

    # sort images bycontrast level
    def tryint(x):
        try:
            return int(x)
        except ValueError:
            return x

    rewardImgs.sort(key=lambda ri: [tryint(c) for c in re.split('([0-9]+)', ri.name)])
    imagePerformance(poke_events, rewardImgs, outputCSV)
    if preset is Presets.SPATIAL or preset is Presets.CONTRAST:
        imagePerformanceFirst(poke_events, rewardImgs, outputCSV)
    print('\n')


def imagePerformance(poke_events, rewardImgs, outputCSV):
    outputCSV.append(["Image Name", "Contrast", "Appearances", "Hits", "Misses", "Success Rate %", "Latency Mean", "Latency SEM"])
    for ri in rewardImgs:
        hits = 0
        for pe in poke_events:
            if pe.image == ri and pe.isSuccess():
                hits += 1  # poke events that occurred in the presence of the REWARD image
        print('REWARD image appearances for {0} >> {1}'.format(ri.name, ri.numAppearances))
        print('Hits/Successful Pokes >> ', hits)
        success_rate = hits * 100.0 / ri.numAppearances if ri.numAppearances else 'N/A'
        outputCSV.append([ri.name, getContrast(ri), ri.numAppearances, hits, ri.numAppearances - hits,
                          success_rate, ri.avg_latency, ri.SEM_latency])


def imagePerformanceFirst(poke_events, rewardImgs, outputCSV):
    outputCSV.append([])
    outputCSV.append(
        ["Image Name", "Contrast", "First Appearances", "Hits", "Misses", "Success Rate %", "Latency Mean", "Latency SEM"])
    for ri in rewardImgs:
        hits = 0
        firstAppearances = 0
        other = 0
        for ap in ri.appearances.values():
            if ap.rewardSeqNum == 1:
                for pe in ap.poke_events:
                    if pe.isSuccess():
                        hits += 1  # poke events that occurred in the presence of the REWARD image
                        break
                firstAppearances += 1

        print('FIRST ONLY REWARD image appearances for {0} >> {1}'.format(ri.name, firstAppearances))
        print('Hits/Successful Pokes >> ', hits)
        print("OTHER >>> ", other)
        success_rate = hits * 100.0 / firstAppearances if firstAppearances else 'N/A'
        outputCSV.append([ri.name, getContrast(ri), firstAppearances, hits, firstAppearances - hits,
                          success_rate, ri.avg_latency_1st, ri.SEM_latency_1st])


def getFileNames(location):
    fileNames = []

    def recursiveDirectories(loc):
        nonlocal fileNames
        try:
            for d in next(os.walk(loc))[1]:
                recursiveDirectories(loc + d + '/')
            for f in next(os.walk(loc))[2]:
                if 'Results' in f and '.txt' in f:
                    fileNames.append(loc + f)
        except StopIteration:
            pass

    recursiveDirectories(location)
    fileNames.sort()
    return fileNames


def initialize(allInput, filename, findFloat):
    images = []
    preset = ''
    mouseNum = 0
    for line in allInput:
        if 'USB drive ID: ' in line:
            print("\n***********************************\n")
            print(filename)
            print(line)
            mouseNum = int(findFloat.search(line).group(0))
        elif 'Control image set:' in line:
            for img in line[line.find('[') + 1:line.rfind(']')].split(','):
                images.append(Image(img.strip(), ImageTypes.CONTROL))
        elif 'Reward image set:' in line:
            for img in line[line.find('[') + 1:line.rfind(']')].split(','):
                images.append(Image(img.strip(), ImageTypes.REWARD))
        elif 'preset: ' in line:
            preset = line.split('preset: ')[1]
            if 'contrast' in preset.lower():
                preset = Presets.CONTRAST
            elif 'spatial' in preset.lower():
                preset = Presets.SPATIAL
            elif '1' in preset:
                preset = Presets.NIGHT_1
            elif '2' in preset:
                preset = Presets.NIGHT_2
            elif '3' in preset:
                preset = Presets.NIGHT_3
            elif '4' in preset:
                preset = Presets.NIGHT_4

        elif "Start of experiment" in line:
            return images, "Mouse_{0}".format(mouseNum), preset


def analyze():
    global LOCALDIR
    if not LOCALDIR.endswith('/'):
        LOCALDIR += '/'
    for filename in getFileNames(LOCALDIR):
        Image.appearanceLog = OrderedDict()  # reset appearances
        with open(filename, 'r') as resultFile:
            allInput = resultFile.readlines()
        findFloat = re.compile("[+-]?([0-9]*[.])?[0-9]+")  # regex to search for a number (float)
        wheelHalfTimes, doorStates, doorTimes, pumpStates, pumpTimes, poke_events, rotation_intervals = [], [], [], [], [], [], []
        skipLine = False
        curImgName = None
        pokeInProgress = False
        images, identifier, preset = initialize(allInput, filename, findFloat)
        images = set(images)  # convert to set to avoid accidental duplication
        Image.images = images

        wb = Workbook()  # open(filename.replace(filename[filename.rfind('/') + 1:], identifier + '.csv'), 'w')
        outputCSV = wb.active
        try:
            controlImgStart = [im for im in images if im.imageType == ImageTypes.CONTROL][0]
        except IndexError:
            print("Warning: No CONTROL Images")
            outputCSV.append(["WARNING: no CONTROL images defined"])
            controlImgStart = [im for im in images][0]
        # ControlImgStart defined in case wheel or door activity is documented prior to first image appearance
        # documentation. This occurs rarely and is a bug in the results file generation protocol.

        currentImg, pokeImg, runImg, currentState = controlImgStart, controlImgStart, controlImgStart, None

        for line in allInput:

            if 'starting' in line:
                continue

            elif 'Image' in line and 'Name:' in line:
                newImgName = line[line.find('Name:') + 5: line.find(',')].strip()
                if curImgName != newImgName:  # ignore if it is the same image (this is a bug)
                    newImg = next((img for img in images if img.name == newImgName), None)
                    assert newImg is not None, 'Unrecognized image: {0}'.format(newImgName)
                    newImg.incrementAppearances(float(re.search("Time: (.*)", line).group(1)), currentImg)
                    curImgName = newImgName
                    currentImg = newImg


            elif 'Wheel' in line and not pokeInProgress:
                if skipLine:
                    skipLine = False
                    continue
                if currentState is Activity.Poking:
                    endPoke(doorStates, doorTimes, pumpTimes, pumpStates, pokeImg, poke_events)
                    doorStates, doorTimes, pumpTimes, pumpStates = [], [], [], []
                currentState = Activity.Running
                if 'State:' in line:
                    wheelHalfTimes.append(float(findFloat.search(line).group(0)))  # appends times
                if 'revolution' in line:
                    # need to skip next data point because wheel state does not actually change; it appears to be a bug
                    skipLine = True
                    continue  # do NOT reset skipLine boolean

            elif 'Pump' in line:
                if re.search("State: (.*), Time", line).group(1) == 'On':
                    pump_state = PumpStates.On
                    pokeImg = currentImg  # the poke event's image should be the image when the pump is on (ie REWARD image)
                    pokeInProgress = True  # ensure parameters don't change within poke duration
                else:
                    pump_state = PumpStates.Off
                    pokeInProgress = False
                pumpStates.append(pump_state)
                pumpTimes.append(float(findFloat.search(line).group(0)))

            elif 'Door' in line:
                if currentState is Activity.Running:
                    endRun(wheelHalfTimes, currentImg, rotation_intervals)
                    wheelHalfTimes = []
                if currentState is not Activity.Poking and not pokeInProgress:
                    pokeImg = currentImg  # record image when poke event starts
                currentState = Activity.Poking
                door_state = DoorStates.High if re.search("State: (.*), Time", line).group(
                    1) == 'High' else DoorStates.Low
                doorStates.append(door_state)
                doorTimes.append(float(findFloat.search(line).group(0)))

            skipLine = False
        if currentState is Activity.Poking:
            endPoke(doorStates, doorTimes, pumpTimes, pumpStates, pokeImg, poke_events)
        else:
            endRun(wheelHalfTimes, currentImg, rotation_intervals)
        pruneRotationIntervals(rotation_intervals)

        analysisFuncs(poke_events, images, filename, wb, preset)
        wb.save(filename.replace(filename[filename.rfind('/') + 1:], identifier + '.xlsx'))


'''ANALYSIS FUNCTION CALLS BEGIN HERE; DO NOT EDIT ABOVE WHEN RUNNING ANALYSIS. CHANGES SHOULD BE MADE ONLY TO 
analysisFuncs METHOD BELOW.'''


def analysisFuncs(poke_events, images, filename, wb, preset):
    outputCSV = wb.active
    setImageLatencies(poke_events, images)  # must be called FIRST
    pokeStatistics(poke_events, images, filename, outputCSV, preset)
    pokesPerHour(poke_events, outputCSV)
    pokeLatencies(wb, preset)
    # outputCSV.close()  # DO NOT delete this line or data may be corrupted (not just results! DATA!!)


analyze()
